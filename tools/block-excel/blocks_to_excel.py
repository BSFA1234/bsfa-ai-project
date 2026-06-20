#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
blocks_to_excel.py
Read extracted SimaticML block XMLs (siblings) and produce an intuitive Excel that
separates FIXED values (identical across all siblings = template / required) from
VARIABLE values (differ per sibling = what changes on copy), and tags each row by
nature: STRUCTURE (network title / instruction order) vs PARAMETER (operands).

Usage:
  python blocks_to_excel.py --in <xml_dir> --out <out.xlsx> [--files OP04.xml OP05.xml ...]

Read-only on the XML. Output Excel contains real operand/symbol names => keep OUT of public repo.
"""
import sys, os, glob, argparse
import xml.etree.ElementTree as ET
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def lname(tag):
    return tag.split('}')[-1]

def text_of(el):
    return (el.text or '').strip() if el is not None else ''

def net_title(cu):
    # network title = MultilingualText CompositionName="Title" -> first non-empty Text
    for mt in cu.iter():
        if lname(mt.tag) == 'MultilingualText' and mt.get('CompositionName') == 'Title':
            for t in mt.iter():
                if lname(t.tag) == 'Text' and (t.text or '').strip():
                    return t.text.strip()
    return ''

def access_str(acc):
    sym = acc.find('{*}Symbol')
    if sym is not None:
        comps = [c.get('Name') for c in sym if lname(c.tag) == 'Component']
        joined = '.'.join([c for c in comps if c])
        if joined:
            return joined
    const = acc.find('{*}Constant')
    if const is not None:
        cn = const.get('Name')
        if cn:
            return 'const:' + cn
        for v in const.iter():
            if lname(v.tag) == 'ConstantValue' and (v.text or '').strip():
                return 'const:' + v.text.strip()
    return '(' + (acc.get('Scope') or '?') + ')'

def parse_block(path):
    root = ET.parse(path).getroot()
    block = None
    for el in root:
        if lname(el.tag).startswith('SW.Blocks.'):
            block = el
            break
    info = {'file': os.path.basename(path), 'name': '?', 'type': '?', 'lang': '?', 'networks': []}
    if block is None:
        return info
    info['type'] = lname(block.tag).split('.')[-1]
    al = block.find('{*}AttributeList')
    if al is not None:
        info['name'] = text_of(al.find('{*}Name')) or info['file']
        info['lang'] = text_of(al.find('{*}ProgrammingLanguage')) or '?'
    ol = block.find('{*}ObjectList')
    if ol is not None:
        idx = 0
        for cu in ol:
            if lname(cu.tag) != 'SW.Blocks.CompileUnit':
                continue
            idx += 1
            net = {'index': idx, 'title': net_title(cu), 'instructions': [], 'operands': []}
            flg = cu.find('.//{*}FlgNet')
            if flg is not None:
                parts = flg.find('{*}Parts')
                if parts is not None:
                    for p in parts:
                        ln = lname(p.tag)
                        if ln == 'Part':
                            name = p.get('Name') or '?'
                            neg = p.find('{*}Negated') is not None
                            net['instructions'].append(name + ('(NC)' if neg else ''))
                        elif ln == 'Call':
                            ci = p.find('{*}CallInfo')
                            net['instructions'].append('Call:' + (ci.get('Name') if ci is not None else '?'))
                        elif ln == 'Access':
                            net['operands'].append(access_str(p))
            info['networks'].append(net)
    return info

# kind -> nature (B classification)
def nature(kind):
    if kind in ('네트워크 제목', '명령 순서'):
        return '구조'
    return '파라미터'

def compare(blocks):
    maxnets = max((len(b['networks']) for b in blocks), default=0)
    fixed, variable = [], []   # fixed: (net, nature, kind, value)  variable: (net, nature, kind, [vals])
    for i in range(maxnets):
        nets = [b['networks'][i] if i < len(b['networks']) else None for b in blocks]
        netno = i + 1
        # title
        titles = [n['title'] if n else '(없음)' for n in nets]
        kind = '네트워크 제목'
        (fixed if len(set(titles)) == 1 else variable).append(
            (netno, nature(kind), kind, titles[0] if len(set(titles)) == 1 else titles))
        # instruction order (whole sequence)
        instr = [' > '.join(n['instructions']) if n and n['instructions'] else '(없음)' for n in nets]
        kind = '명령 순서'
        (fixed if len(set(instr)) == 1 else variable).append(
            (netno, nature(kind), kind, instr[0] if len(set(instr)) == 1 else instr))
        # operands, aligned by position
        maxops = max((len(n['operands']) if n else 0 for n in nets), default=0)
        for j in range(maxops):
            vals = [(n['operands'][j] if n and j < len(n['operands']) else '(없음)') for n in nets]
            kind = '피연산자 #%d' % (j + 1)
            (fixed if len(set(vals)) == 1 else variable).append(
                (netno, nature(kind), kind, vals[0] if len(set(vals)) == 1 else vals))
    return fixed, variable

# ---- styling helpers ----
HEAD = PatternFill('solid', fgColor='1F4E78')
HEADF = Font(bold=True, color='FFFFFF')
FIXFILL = PatternFill('solid', fgColor='E2EFDA')     # light green
VARFILL = PatternFill('solid', fgColor='FCE4D6')     # light orange
STRUCTF = Font(color='2E75B6', bold=True)
PARAMF = Font(color='C55A11', bold=True)
THIN = Side(style='thin', color='BFBFBF')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical='top')
CTR = Alignment(horizontal='center', vertical='center')

def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEAD; cell.font = HEADF; cell.alignment = CTR; cell.border = BORDER

def nature_cell(cell):
    cell.value = cell.value
    if cell.value == '구조':
        cell.font = STRUCTF
    elif cell.value == '파라미터':
        cell.font = PARAMF
    cell.alignment = CTR

def build_excel(blocks, fixed, variable, out):
    wb = openpyxl.Workbook()

    # --- Sheet 1: 개요 ---
    ws = wb.active; ws.title = '개요'
    ws.append(['블록명', '종류', '언어', '네트워크 수'])
    style_header(ws, 4)
    for b in blocks:
        ws.append([b['name'], b['type'], b['lang'], len(b['networks'])])
    ws.append([])
    ws.append(['범례'])
    ws.append(['고정값(공통)', '= 모든 형제가 동일한 값 (템플릿/필수 — 복붙 시 그대로 둬야 함)'])
    ws.append(['변동값(형제별)', '= 형제마다 다른 값 (복붙 시 바꿔야 하는 값)'])
    ws.append(['성격: 구조', '= 네트워크 제목 / 명령 순서 (로직 뼈대)'])
    ws.append(['성격: 파라미터', '= 피연산자 / 주소 / DB참조 (채워 넣는 값)'])
    ws.append(['정렬 기준', '네트워크는 위치(순서)로 형제끼리 맞춤. 구조가 다르면 변동으로 표시됨.'])
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 70
    for col in 'CD':
        ws.column_dimensions[col].width = 12

    # --- Sheet 2: 고정값(공통) ---
    ws = wb.create_sheet('고정값(공통)')
    ws.append(['네트워크', '성격', '항목', '값 (모든 형제 동일)'])
    style_header(ws, 4)
    for (netno, nat, kind, val) in fixed:
        ws.append([netno, nat, kind, val])
        r = ws.max_row
        nature_cell(ws.cell(row=r, column=2))
        for c in range(1, 5):
            ws.cell(row=r, column=c).border = BORDER
            ws.cell(row=r, column=c).alignment = WRAP if c == 4 else CTR
        ws.cell(row=r, column=1).fill = FIXFILL
    ws.freeze_panes = 'A2'
    ws.column_dimensions['A'].width = 10; ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 16; ws.column_dimensions['D'].width = 80

    # --- Sheet 3: 변동값(형제별) ---
    ws = wb.create_sheet('변동값(형제별)')
    names = [b['name'] for b in blocks]
    ws.append(['네트워크', '성격', '항목'] + names)
    style_header(ws, 3 + len(names))
    for (netno, nat, kind, vals) in variable:
        ws.append([netno, nat, kind] + list(vals))
        r = ws.max_row
        nature_cell(ws.cell(row=r, column=2))
        for c in range(1, 4 + len(names)):
            ws.cell(row=r, column=c).border = BORDER
            ws.cell(row=r, column=c).alignment = WRAP if c >= 4 else CTR
        for c in range(4, 4 + len(names)):
            ws.cell(row=r, column=c).fill = VARFILL
    ws.freeze_panes = 'D2'
    ws.column_dimensions['A'].width = 10; ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 16
    for i in range(len(names)):
        ws.column_dimensions[chr(ord('D') + i)].width = 28

    # --- Sheet 4: 성격요약 (B) ---
    ws = wb.create_sheet('성격요약(구조vs파라미터)')
    ws.append(['성격', '고정(공통)', '변동(형제별)', '합계'])
    style_header(ws, 4)
    def cnt(rows, nat): return sum(1 for r in rows if r[1] == nat)
    for nat in ['구조', '파라미터']:
        f = cnt(fixed, nat); v = cnt(variable, nat)
        ws.append([nat, f, v, f + v])
        r = ws.max_row; nature_cell(ws.cell(row=r, column=1))
        for c in range(1, 5): ws.cell(row=r, column=c).border = BORDER
    ws.append(['합계', len(fixed), len(variable), len(fixed) + len(variable)])
    for c in range(1, 5):
        ws.cell(row=ws.max_row, column=c).font = Font(bold=True); ws.cell(row=ws.max_row, column=c).border = BORDER
    for col, w in zip('ABCD', [12, 12, 14, 10]):
        ws.column_dimensions[col].width = w

    wb.save(out)
    return out

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--in', dest='indir', required=True)
    ap.add_argument('--out', dest='out', required=True)
    ap.add_argument('--files', nargs='*', default=None)
    a = ap.parse_args()
    if a.files:
        paths = [os.path.join(a.indir, f) for f in a.files]
    else:
        paths = sorted(glob.glob(os.path.join(a.indir, '*.xml')))
    if not paths:
        print('no xml found in', a.indir); sys.exit(1)
    blocks = [parse_block(p) for p in paths]
    blocks.sort(key=lambda b: b['name'])
    fixed, variable = compare(blocks)
    out = build_excel(blocks, fixed, variable, a.out)
    print('blocks:', ', '.join(b['name'] for b in blocks))
    print('fixed rows:', len(fixed), ' variable rows:', len(variable))
    print('saved:', out)

if __name__ == '__main__':
    main()
